import matplotlib.pyplot as plt
import datetime as dt
import numpy as np
import pandas as pd
import sys
import math
from openpyxl import load_workbook
import os

error_dict = {'case': [], '가열로 호기': [], '시간': [], '온도': [], '가스': []}
case_id = 0


class HF:
    index = 0
    index_h = 0
    index_h_next = 0
    flag_end = 0
    temp_1 = 0
    temp_2 = 0
    temp_list = []
    temp_in = []
    temp_out = []
    temp_time = dt.timedelta()
    df = pd.DataFrame()

    def __init__(self):
        self.index = 0
        self.df = pd.DataFrame(columns=['가열로 번호', '시작시간', '종료시간', '가스사용량', 'Type', '소재 list',
                                        '작업일자', '주/야간', 'in', 'out', 'cycle', '평일/주말'])

    def in_time(self, df_mat):
        self.df = self.df.sort_values(["시작시간"], ascending=[True])
        self.df = self.df.reset_index(drop=True)
        for i in range(len(self.df.index)):
            if self.df['Type'].loc[i] == 'heat':
                temp = []
                mass = []
                for j in range(len(df_mat.index)):
                    num = '가열로' + str(self.df['가열로 번호'].loc[i]) + '호기'
                    if num == df_mat['장입가열로'].loc[j] and \
                            str(self.df['실제 시작시간'].loc[i]) == str(df_mat['시작시간'].loc[j]) + ':00':
                        if pd.isna(df_mat.loc[j, '시리얼번호']) or len(df_mat.loc[j, '시리얼번호'].split(',')) > int(
                                df_mat.loc[j, '투입수량']):
                            self.df.loc[i, 'drop_flag'] = 1
                        elif len(df_mat.loc[j, '시리얼번호'].split(',')) <= int(df_mat.loc[j, '투입수량']):
                            for k in range(len(df_mat.loc[j, '시리얼번호'].split(','))):
                                temp.append(df_mat['수주번호'].loc[j])
                self.df['소재 list'].loc[i] = temp

    def sett(self, df, s='./HF_out/test_2019_'):
        self.in_time(df)
        self.out(s)

    def out(self, s='./HF_out/test_2019_'):
        self.df.to_csv(s + str(self.df['가열로 번호'].loc[2]) + ".csv", mode='w', encoding='euc-kr')

    def set_next_h(self):
        self.index_h_next = self.index_h
        temp = self.index_h_next
        if self.flag_end == 0:
            while 1:
                if temp < len(self.df.index) - 1:
                    temp += 1
                    if self.df['Type'].loc[temp] == 'heat':
                        break
                    else:
                        pass
                if temp == len(self.df.index) - 1:
                    self.flag_end = 1
                    break
            self.index_h_next = temp
        else:
            pass

    def search_next_h(self):
        temp = self.index_h_next
        if self.flag_end == 0:
            while 1:
                if temp < len(self.df.index) - 1:
                    temp += 1
                    if self.df['Type'].loc[temp] == 'heat':
                        break
                    else:
                        pass
                if temp == len(self.df.index) - 1:
                    self.flag_end = 1
                    break
            self.index_h = self.index_h_next
            self.index_h_next = temp
            self.index = self.index_h
            self.temp_1 = self.index_h
            self.temp_2 = self.index_h
            self.temp_list = self.df['소재 list'].loc[self.index_h][:]
        else:
            pass

    def change_list(self):
        for i in range(len(self.df.index)):
            temp = []
            temp2 = []
            if self.df['Type'].loc[i] == 'heat' and type(self.df['소재 list'].loc[i]) != float:
                temp = (self.df['소재 list'].loc[i][1:len(self.df['소재 list'].loc[i]) - 1]).split(',')
                for k in temp:
                    # print(k.strip(), k.strip()[1:len(k.strip())-1])
                    temp2.append(k.strip()[1:len(k.strip()) - 1])
                self.df['소재 list'].loc[i] = temp2[:]
        self.temp_list = self.df['소재 list'].loc[self.index_h][:]

    def change_list2(self):
        for i in range(len(self.df.index)):
            temp = []
            temp2 = []
            temp_1 = []
            temp2_1 = []
            temp_2 = []
            temp2_2 = []
            if type(self.df['소재 list'].loc[i]) != float:
                temp = (self.df['소재 list'].loc[i][1:len(self.df['소재 list'].loc[i]) - 1]).split(',')
                # temp_1 = (self.df['out'].loc[i][1:len(self.df['out'].loc[i]) - 1]).split(',')
                # temp_2 = (self.df['in'].loc[i][1:len(self.df['in'].loc[i]) - 1]).split(',')
                for k in temp:
                    # print(k.strip(), k.strip()[1:len(k.strip())-1])
                    temp2.append(k.strip()[1:len(k.strip()) - 1])
                for k_1 in temp_1:
                    # print(k.strip(), k.strip()[1:len(k.strip())-1])
                    temp2_1.append(k_1.strip()[1:len(k_1.strip()) - 1])
                for k_2 in temp_2:
                    # print(k.strip(), k.strip()[1:len(k.strip())-1])
                    temp2_2.append(k_2.strip()[1:len(k_2.strip()) - 1])
                self.df['소재 list'].loc[i] = temp2[:]
                # self.df['out'].loc[i] = temp2_1[:]
                # self.df['in'].loc[i] = temp2_2[:]
            # print(i)

    def fill(self):
        count_c = 0
        # print(len(self.df.index))
        for i, row in self.df.iterrows():
            # print(i)
            if pd.isna(self.df.loc[i, '소재 list']) and i > 0:
                if self.df.loc[i, 'Type'] == 'open' and i > 0:
                    self.df.loc[i, '소재 list'] = self.df.loc[i-1, '소재 list']
                elif self.df.loc[i, 'Type'] == 'hold' and i > 0:
                    self.df.loc[i, '소재 list'] = self.df.loc[i - 1, '소재 list']
                elif self.df.loc[i, 'Type'] == 'reheat' and i > 0:
                    self.df.loc[i, '소재 list'] = self.df.loc[i - 1, '소재 list']
            elif not pd.isna(self.df.loc[i, '소재 list']) and i > 0:
                if self.df.loc[i, 'Type'] == 'hold' and i > 0:
                    self.df.loc[i, '소재 list'] = self.df.loc[i - 1, '소재 list']
                elif self.df.loc[i, 'Type'] == 'reheat' and i > 0:
                    self.df.loc[i, '소재 list'] = self.df.loc[i - 1, '소재 list']
            if self.df.loc[i, 'Type'] == 'heat':
                count_c += 1
            self.df['cycle'].loc[i] = count_c
            # if self.df['Type'].loc[i] == 'reheat' and self.df['in'].loc[i - 1] != '[]':
            #     self.df['in'].loc[i] = self.df['in'].loc[i - 1]
            if self.df['Type'].loc[i] != 'heat':
                self.df['작업일자'].loc[i] = self.df['작업일자'].loc[i - 1]
                self.df['주/야간'].loc[i] = self.df['주/야간'].loc[i - 1]

    def set_cycle(self):
        count_c = 0
        for i in range(len(self.df.index)):
            if self.df['Type'].loc[i] == 'heat':
                count_c += 1
            self.df['cycle'].loc[i] = count_c

    def match_time(self, df_t):
        for i in range(len(self.df.index)):
            # d1 = dt.datetime.strptime(self.df['시작시간1'].loc[i], "%Y-%m-%d %H:%M:%S")
            if self.df['Type'].loc[i] == 'heat':
                d1 = dt.datetime.strptime(self.df['실제 시작시간'].loc[i], "%Y-%m-%d %H:%M:%S")
                for j in range(len(df_t.index)):
                    d2 = dt.datetime.strptime(df_t['가열시작일시'].loc[j], "%Y-%m-%d %H:%M")
                    num = '가열로' + str(self.df['가열로 번호'].loc[i]) + '호기'
                    if num == df_t['가열로명'].loc[j] and \
                            d1 == d2:
                        # print('work', df_t['작업일자'].loc[j], df_t['주/야간'].loc[j])
                        self.df['작업일자'].loc[i] = df_t['작업일자'].loc[j]
                        self.df['주/야간'].loc[i] = df_t['주/야간'].loc[j]
                        break

    def week(self):
        to_day = ['월', '화', '수', '목', '금', '토', '일']
        for i, row in self.df.iterrows():
            if self.df.loc[i, 'Type'] == 'heat':
                time = dt.datetime.strptime(self.df.loc[i, '시작시간'], "%Y-%m-%d %H:%M:%S")
                day_ = to_day[time.weekday()]
                if day_ == '토' or day_ == '일':
                    self.df.loc[i, '평일/주말'] = '주말'
                elif day_ == '월' and i > 0:
                    time_past = dt.datetime.strptime(self.df.loc[i - 1, '시작시간'], "%Y-%m-%d %H:%M:%S")
                    day_past = to_day[time_past.weekday()]
                    if day_past == '토' or day_past == '일':
                        self.df.loc[i, '평일/주말'] = '평일'
                    else:
                        self.df.loc[i, '평일/주말'] = '주말'
                else:
                    self.df.loc[i, '평일/주말'] = '평일'


def get_data_excel(data, s, num):
    global error_dict
    global case_id
    ex = load_workbook(s, data_only=True)['이력']
    MM = ex.max_row
    for i in range(2, MM):
        # print(i)
        temp_time = dt.datetime.strptime(ex['B' + str(i)].value, "%Y-%m-%d %H:%M:%S")
        # print(ex['B' + str(i)].value, 'row :', i)
        # print(ex['C' + str(i)].value, type(ex['C' + str(i)].value))
        if str(ex['C' + str(i)].value) == '-' or int(ex['C' + str(i)].value) == 0:
            # print('gas off in')
            temp_tem = 0
            temp_off = 1
        else:
            # print('gas right in')
            temp_tem = float(ex['C' + str(i)].value)
            temp_off = 0
        # print(ex['D' + str(i)].value, type(ex['D' + str(i)].value), ex['D' + str(i)].value == '-')
        if str(ex['D' + str(i)].value) == '-' or float(ex['D' + str(i)].value) < 0:
            # print('tmpe off in')
            temp_gas = 0
            gas_off = 1
        else:
            if i < MM and str(ex['D' + str(i + 1)].value) != '-':
                if float(ex['D' + str(i + 1)].value) > 200:
                    print('error!! : ', temp_time, float(ex['D' + str(i + 1)].value))
                    for k in range(-2, 5):
                        error_dict['case'].append(case_id)
                        error_dict['가열로 호기'].append(num)
                        error_dict['시간'].append(ex['B' + str(i + k)].value)
                        error_dict['온도'].append(ex['C' + str(i + k)].value)
                        error_dict['가스'].append(ex['D' + str(i + k)].value)
                    temp_gas = 0
                    gas_off = 1
                    case_id += 1
                else:
                    temp_gas = float(ex['D' + str(i)].value)
                    gas_off = 0
            else:
                temp_gas = float(ex['D' + str(i)].value)
                gas_off = 0
        data.append(
            {'TEMPERATURE': temp_tem, 'GAS': temp_gas, 'TIME': temp_time, 'GAS_OFF': gas_off, 'TEMP_OFF': temp_off})


def wrong_st_ed(s, work):
    df = pd.read_csv(s + '.csv')
    df = df.fillna(0)
    df_temp = pd.DataFrame()
    del_arr = []
    count0 = 0
    for num in work:
        del_arr2 = []
        df_new = pd.DataFrame()
        count = 0
        for i, row in df.iterrows():
            if df['가열로명'].loc[i] == '가열로' + str(num) + '호기':
                df_new = df_new.append(row)
                df_new = df_new.reset_index(drop=True)
        print(len(df_new))
        count0 += len(df_new)
        if len(df_new.index) != 0:
            df_new = df_new.sort_values(['가열시작일시'], ascending=[True])
            df_new = df_new.reset_index(drop=True)
            for j, row in df_new.iterrows():
                if j > 0:
                    start = dt.datetime.strptime(df_new['가열시작일시'].loc[j], "%Y-%m-%d %H:%M")
                    end2 = dt.datetime.strptime(df_new['가열종료일시'].loc[j], "%Y-%m-%d %H:%M")
                    end = dt.datetime.strptime(df_new['가열종료일시'].loc[j - 1], "%Y-%m-%d %H:%M")
                    if (start - end).total_seconds() < 0:
                        # print(int(df_new['순번'].loc[j]), int(df_new['순번'].loc[j-1]))
                        print(start, end, start - end, j)
                        del_arr.append(int(df_new['순번'].loc[j - 1]))
                        del_arr.append(int(df_new['순번'].loc[j]))
                        del_arr2.append(j - 1)
                        del_arr2.append(j)
                        count += 2
                    if (end2 - start).total_seconds() < 0:
                        del_arr.append(int(df_new['순번'].loc[j]))
                        del_arr2.append(j)
                        count += 0.01
            print('가열로 ' + str(num) + '호기 : ', count)
            print(del_arr2)
            del_arr2 = set(del_arr2)
            for t in del_arr2:
                df_new = df_new.drop([t])
            df_temp = pd.concat([df_temp, df_new])
            df_temp = df_temp.reset_index(drop=True)
    df_temp.to_csv(s + '_re1.csv', encoding='euc-kr')


# 시작시간 - 종료시간 read
def st_end_all(num, time):
    start_arr = []
    end_arr = []
    flag = 0
    df_st_end_sig = pd.read_csv(time, encoding='euc-kr')
    df_st_end_sig = df_st_end_sig.fillna(0)
    # print('가열로' + str(num) + '호기')
    for i in range(len(df_st_end_sig.index)):
        if df_st_end_sig['가열로명'].loc[i] == '가열로' + str(num) + '호기':
            # print(i, df_st_end_sig['가열시작일시'].loc[i], df_st_end_sig['가열종료일시'].loc[i])
            if df_st_end_sig['가열시작일시'].loc[i] != 0 and df_st_end_sig['가열종료일시'].loc[i] != 0:
                start = dt.datetime.strptime(df_st_end_sig['가열시작일시'].loc[i], "%Y-%m-%d %H:%M")
                end = dt.datetime.strptime(df_st_end_sig['가열종료일시'].loc[i], "%Y-%m-%d %H:%M")
                for s in start_arr:
                    if start == s:
                        flag = 1
                        print(start)
                if flag == 0:
                    start_arr.append(start)
                flag = 0
                for e in end_arr:
                    if end == e:
                        flag = 1
                        print(end)
                if flag == 0:
                    end_arr.append(end)
                flag = 0
    return start_arr, end_arr


# Reinforcement
def reinforce(data):
    flag = 0
    temp1 = 0
    temp2 = 0
    num = 0  # 횟수
    temp3 = 0
    gra = 0
    temp_min = 0  # 최대시간

    for i in range(len(data)):

        if float(data[i]['TEMPERATURE']) > 2000:
            data[i]['TEMPERATURE'] = 0

        if (flag == 0 and temp1 == 0 and float(data[i]['TEMPERATURE']) == 0):
            temp1 = i
            temp2 = float(data[i - 1]['TEMPERATURE'])
            flag = 1
        elif (float(data[i]['TEMPERATURE']) != 0 and flag == 1):
            temp3 = float(data[i]['TEMPERATURE'])
            gra = (temp3 - temp2) / (i - temp1)

            for j in range(temp1, i):
                data[j]['TEMPERATURE'] = float(data[j - 1]['TEMPERATURE']) + gra
                temp_min += 1

            flag = 0
            temp1 = 0
            temp2 = 0
            temp3 = 0
            gra = 0
            num += 1

    # print(temp_min)
    # print(num)


# Data plot
def plotting(data, change_point10, start_arr, end_arr, num, re, start_real=None, end_real=None):
    x = []
    y = []
    z = []
    xx = []
    plt.figure()
    for i in range(len(data)):
        x.append(data[i]['TIME'])
        y.append(data[i]['TEMPERATURE'])
        z.append(data[i]['GAS'] * 5)
    plt.title(str(num) + "heat")
    plt.ylabel("temperature")

    # 가스, 온도
    plt.plot(x, y, color='dimgrey')
    plt.plot(x, z, color='black')

    # change point / 수직선들

    if start_real is not None:
        for i in start_real:
            plt.axvline(x=i, color='orange')

    if end_real is not None:
        for i in end_real:
            plt.axvline(x=i, color='lime')

    '''
    for i in range(len(x)):
        if change_point10[i] is not None:
            xx.append(x[i])
    for i in xx:
        plt.axvline(x=i)
    '''

    for i in re:
        plt.axvline(x=i, color='purple')
    for i in start_arr:
        plt.axvline(x=i, color='red')
    for i in end_arr:
        plt.axvline(x=i, color='green')
    # for i in change_point10:
    #     plt.axvline(x=i, color='blue')
    '''
    if end_real is not None:
        for i in end_real:
            plt.axvline(x=i, color='green')
    '''
    plt.plot(x, change_point10, 'o', color='red')

    '''
    data_Sdeviation.clear()
    count=0
    max = 0
    max_count = 0
    flag = 0
    temp=0
    for i in range(len(data)):
        data_Sdeviation.append(None)
        if float(data[i]) == 0:
            count += 1
            max_count += 1
            data_Sdeviation[i] = 0
            temp=i
            flag = 1
            #print(i,max_count)
        elif flag == 1 and float(data[i]) != 0:
            if max_count > max:
                max = max_count
                print(i, max)
            max_count=0
            flag=0
    print(count,temp,max)
    plt.plot(data_Sdeviation, 'o', color='red', lw=5)
    '''
    # plt.plot(data_S, 'o', color='yellow')
    # plt.show()


def plotting2(data, start_arr, end_arr, num):
    x = []
    y = []
    z = []
    for i in range(len(data)):
        x.append(data[i]['TIME'])
        y.append(data[i]['TEMPERATURE'])
        z.append(data[i]['GAS'] * 10)
    plt.figure()
    plt.xlabel("time")
    plt.ylabel("temperature")
    plt.plot(x, y)
    for i in start_arr:
        plt.axvline(x=i, color='red')
    for i in end_arr:
        plt.axvline(x=i, color='green')
    plt.plot(x, z, color='black')
