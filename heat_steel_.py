import pandas as pd
import os
from openpyxl import load_workbook


def excel_test(s, s2, df):
    ex1 = load_workbook(s, data_only=True)['생산계획']
    index = 9
    # print(ex1.max_row, type(ex1.max_row))
    for i in range(index, ex1.max_row, 2):
        # print(i, len((ex1['AR' + str(i)].value).split('\n')), (ex1['AR' + str(i)].value).split('\n'))
        # print(i, ex1['AR' + str(i)].value)
        second_index = 2
        if ex1['AR' + str(i)].value is None or ex1['AR' + str(i)].value == '-' or type(ex1['AR' + str(i)].value) != str:
            pass
        else:
            # print()
            # print(i)
            # print(ex1['AR' + str(i)].value)
            # print('in')
            arr1 = (ex1['AR' + str(i)].value).split('\n')
            arr2 = []
            flag = 0
            while flag == 0:
                second_index += 2
                if ex1['AR' + str(i + second_index)].value is not None:
                    flag = 1
                if (i + second_index) == ex1.max_row:
                    flag = 1
                # print('second index : ', second_index)
                # print(ex1['AR' + str(i + second_index)].value)
            # print('second index : ', second_index)
            if len(arr1) > 1:
                # print(arr1)
                for t in range(len(arr1)):
                    if len(arr1[t].split(' ')) > 1:
                        arr1[t] = arr1[t].split(' ')[0]
                for i2 in range(1, second_index, 2):
                    if ex1['AS' + str(i + i2)].value is not None:
                        excel_test2(ex1, i, i2, arr2)
            elif len((ex1['AR' + str(i)].value).split('\n')) == 1:
                # print(arr1)
                if arr1[0] == '냉괴이송':
                    pass
                else:
                    for i2 in range(1, second_index, 2):
                        if ex1['AS' + str(i + i2)].value is not None:
                            excel_test2(ex1, i, i2, arr2)
            if len(arr2) != 0:
                # pass
                # print(i)
                # print(arr1)
                # print(arr2)
                for t in arr2:
                    # print(t)
                    df = df.append([[t[0], t[1], t[2], s2]])
                    df = df.reset_index(drop=True)
    return df


def excel_test2(ex1, i, i2, arr2):
    index_front = 0
    index_back = len(ex1['AS' + str(i + i2)].value)
    index_front += ex1['AS' + str(i + i2)].value.count('(')
    index_back -= ex1['AS' + str(i + i2)].value.count(')')
    if len(ex1['AS' + str(i + i2)].value[index_front:index_back].split(',')) > 1:
        # print('i값 : ' + str(i) + ' / i2값 : ' + str(i2) + ' / i + i2 :' + str(i + i2))
        # print('원본 : ' + ex1['AS' + str(i + i2)].value[index_front:index_back])
        for t in ex1['AS' + str(i + i2)].value[index_front:index_back].split(', '):
            if len(t.split(',\n')) > 1:
                for k in t.split(',\n'):
                    # print(k)
                    arr2.append([k.strip(), ex1['AP' + str(i + i2 - 1)].value, [ex1['AT' + str(i + i2 -1)].value, ex1['AT' + str(i + i2)].value]])
            elif len(t.split('\n')) > 1:
                for k in t.split('\n'):
                    if len(k) < 1:
                        pass
                    else:
                        # print(k)
                        arr2.append([k.strip(), ex1['AP' + str(i + i2 - 1)].value, [ex1['AT' + str(i + i2 -1)].value, ex1['AT' + str(i + i2)].value]])
            elif len(t.split(' \n')) > 1:
                for k in t.split(' \n'):
                    if len(k) < 1:
                        pass
                    else:
                        # print(k)
                        arr2.append([k.strip(), ex1['AP' + str(i + i2 - 1)].value, [ex1['AT' + str(i + i2 -1)].value, ex1['AT' + str(i + i2)].value]])
            else:
                # print(t)
                arr2.append([t.strip(), ex1['AP' + str(i + i2 - 1)].value, [ex1['AT' + str(i + i2 -1)].value, ex1['AT' + str(i + i2 )].value]])
    else:
        # print('i값 : ' + str(i) + ' / i2값 : ' + str(i2) + ' / i + i2 :' + str(i + i2))
        # print('원본 : ' + ex1['AS' + str(i + i2)].value[index_front:index_back])
        # print(ex1['AS' + str(i + i2)].value[index_front:index_back])
        arr2.append([ex1['AS' + str(i + i2)].value[index_front:index_back], ex1['AP' + str(i + i2 - 1)].value, [ex1['AT' + str(i + i2 -1)].value, ex1['AT' + str(i + i2)].value]])


kkk = ['2월 2주차']
df = pd.DataFrame()
for t in os.listdir('./data_201907~202003/제강생산계획/'):
# for t in kkk:
    path = './data_201907~202003/제강생산계획/' + t
    print(path)
    index = -1
    while os.listdir(path)[index][0] != t[0] or\
            os.listdir(path)[index][-4:-1] != 'xls':
        # print('work')2
        index -= 1
    print(path + '/' + os.listdir(path)[index])
    df = excel_test(path + '/' + os.listdir(path)[index], t, df)
df.to_csv('./data_201907~202003/heat_steel.csv', encoding='euc-kr')